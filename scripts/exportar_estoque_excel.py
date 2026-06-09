#!/usr/bin/env python3
"""
Exporta produtos com estoque do Bling ERP para Excel.

Uso normal (token salvo no banco):
  python3 scripts/exportar_estoque_excel.py

Se o refresh_token expirou, passe um access_token diretamente:
  python3 scripts/exportar_estoque_excel.py --token SEU_ACCESS_TOKEN_AQUI

Como obter um access_token temporário no Bling:
  1. Abra https://developer.bling.com.br → Meus Apps → selecione o app
  2. Clique em "Testar API" ou use o fluxo OAuth normalmente pelo BibelôCRM:
     docker compose up -d api redis  (no /opt/pessoal/bibelocrm)
     Acesse https://crm.papelariabibelo.com.br → Configurações → Bling → Conectar
  3. Copie o access_token gerado e passe via --token
"""

import argparse
import base64
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import psycopg2
import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Configuração ───────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
ENV_FILE   = SCRIPT_DIR.parent / ".env"
BLING_API  = "https://api.bling.com.br/Api/v3"

# Limitar requisições a ~3/s para respeitar rate limit do Bling
RATE_LIMIT_DELAY = 0.35


# ── Leitura do .env ────────────────────────────────────────────────────────────

def load_env(path: Path) -> dict:
    env = {}
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


# ── Tokens Bling ───────────────────────────────────────────────────────────────

def get_db_conn(env: dict):
    # Tenta primeiro o IP interno do container Docker (quando rodando no host)
    # Depois tenta localhost com porta exposta (se configurado)
    for host, port in [("172.21.0.2", 5432), ("localhost", int(env.get("DB_EXPOSED_PORT", 5433)))]:
        try:
            return psycopg2.connect(
                host=host,
                port=port,
                dbname=env.get("DB_NAME", "bibelocrm"),
                user=env.get("DB_USER", "bibelocrm"),
                password=env.get("DB_PASS", ""),
                connect_timeout=5,
            )
        except psycopg2.OperationalError:
            continue
    raise RuntimeError("Não foi possível conectar ao banco bibelocrm")


def get_stored_tokens(conn) -> dict | None:
    with conn.cursor() as cur:
        cur.execute("SELECT ultimo_id FROM sync.sync_state WHERE fonte = 'bling'")
        row = cur.fetchone()
    if not row or not row[0]:
        return None
    try:
        return json.loads(row[0])
    except Exception:
        return None


def save_tokens(conn, tokens: dict) -> None:
    with conn.cursor() as cur:
        cur.execute(
            "UPDATE sync.sync_state SET ultimo_id = %s, ultima_sync = NOW() WHERE fonte = 'bling'",
            [json.dumps(tokens)],
        )
    conn.commit()


def refresh_bling_token(env: dict, refresh_token: str) -> dict:
    client_id     = env["BLING_CLIENT_ID"]
    client_secret = env["BLING_CLIENT_SECRET"]
    credentials   = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()

    print("  Renovando token Bling...", end=" ", flush=True)
    resp = requests.post(
        f"{BLING_API}/oauth/token",
        data={"grant_type": "refresh_token", "refresh_token": refresh_token},
        headers={
            "Authorization": f"Basic {credentials}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        timeout=20,
    )
    if not resp.ok:
        print(f"\n❌  Erro Bling ({resp.status_code}): {resp.text[:300]}")
        resp.raise_for_status()
    data = resp.json()
    tokens = {
        "access_token":  data["access_token"],
        "refresh_token": data["refresh_token"],
        "expires_in":    data["expires_in"],
        "expires_at":    datetime.fromtimestamp(
            time.time() + data["expires_in"], tz=timezone.utc
        ).isoformat(),
    }
    print("ok ✓")
    return tokens


def get_valid_token(env: dict, conn) -> str:
    stored = get_stored_tokens(conn)
    if not stored:
        sys.exit("❌  Nenhum token Bling no banco. Execute o fluxo OAuth primeiro.")

    expires_at = datetime.fromisoformat(stored["expires_at"])
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)

    if time.time() > expires_at.timestamp() - 300:
        tokens = refresh_bling_token(env, stored["refresh_token"])
        save_tokens(conn, tokens)
        return tokens["access_token"]

    return stored["access_token"]


# ── Chamadas Bling ─────────────────────────────────────────────────────────────

def bling_get(url: str, token: str, params: dict | None = None) -> dict:
    time.sleep(RATE_LIMIT_DELAY)
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params=params,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def fetch_all_products(token: str) -> list[dict]:
    """Busca todos os produtos ativos paginados."""
    products = []
    page = 1
    while True:
        print(f"  Produtos — página {page}...", end="\r", flush=True)
        data = bling_get(f"{BLING_API}/produtos", token, {"pagina": page, "limite": 100})
        items = data.get("data", [])
        if not items:
            break
        products.extend(items)
        page += 1
    print(f"  Produtos — {len(products)} encontrados          ")
    return products


def fetch_stock_for_ids(product_ids: list[str], token: str) -> dict[str, float]:
    """Retorna {bling_id: saldo_fisico_total} em lotes de 50."""
    BATCH = 50
    saldos: dict[str, float] = {}
    total_batches = (len(product_ids) + BATCH - 1) // BATCH

    for i in range(0, len(product_ids), BATCH):
        batch = product_ids[i : i + BATCH]
        batch_num = i // BATCH + 1
        print(f"  Estoque — lote {batch_num}/{total_batches}...", end="\r", flush=True)

        params_str = "&".join(f"idsProdutos[]={pid}" for pid in batch)
        time.sleep(RATE_LIMIT_DELAY)
        resp = requests.get(
            f"{BLING_API}/estoques/saldos?{params_str}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        for entry in data.get("data", []):
            produto = entry.get("produto") or {}
            pid     = str(produto.get("id", ""))
            if not pid:
                continue
            saldo = float(entry.get("saldoFisicoTotal") or 0)
            saldos[pid] = saldos.get(pid, 0) + saldo

    print(f"  Estoque — {len(saldos)} registros obtidos          ")
    return saldos


def fetch_categories(token: str) -> dict[int, str]:
    """Retorna {bling_category_id: nome}."""
    categories: dict[int, str] = {}
    page = 1
    while True:
        data = bling_get(f"{BLING_API}/categorias/produtos", token, {"pagina": page, "limite": 100})
        items = data.get("data", [])
        if not items:
            break
        for cat in items:
            categories[cat["id"]] = cat.get("descricao") or cat.get("nome") or "—"
        page += 1
    return categories


# ── Geração do Excel ───────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1A2B4A")
ALT_FILL      = PatternFill("solid", fgColor="F0F4FA")
YELLOW_FILL   = PatternFill("solid", fgColor="FFF9C4")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT     = Font(bold=True)
CURRENCY_FMT  = 'R$\\ #,##0.00'
QTY_FMT       = '#,##0.##'

THIN = Side(style="thin", color="D0D7E3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Código",           16, None),
    ("Nome do Produto",  48, None),
    ("Categoria",        24, None),
    ("Tipo",              8, None),
    ("Unidade",          10, None),
    ("Preço Custo (R$)", 18, CURRENCY_FMT),
    ("Preço Venda (R$)", 18, CURRENCY_FMT),
    ("Estoque Bling",    16, QTY_FMT),
    ("Contagem Física",  16, QTY_FMT),   # ← usuário preenche
    ("Diferença",        14, QTY_FMT),   # ← fórmula automática
]


def build_excel(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário Bibelô"
    ws.freeze_panes = "A2"

    # ── Cabeçalho ────────────────────────────────────────────────
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # ── Dados ─────────────────────────────────────────────────────
    for row_idx, item in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        data_row = [
            item.get("codigo") or "",
            item.get("nome") or "Sem nome",
            item.get("categoria") or "—",
            item.get("tipo") or "P",
            item.get("unidade") or "UN",
            item.get("preco_custo") or 0,
            item.get("preco_venda") or 0,
            item.get("saldo_fisico") or 0,
            None,   # Contagem Física — usuário preenche
            None,   # Diferença — fórmula
        ]
        for col_idx, value in enumerate(data_row, start=1):
            _, _, fmt = COLUMNS[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill.fgColor and fill.fgColor.rgb != "00000000":
                cell.fill = fill
            if fmt:
                cell.number_format = fmt

        # Coluna Contagem Física — fundo amarelo (área editável)
        contagem_cell = ws.cell(row=row_idx, column=9)
        contagem_cell.fill = YELLOW_FILL
        contagem_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Coluna Diferença — fórmula  =I{n}-H{n}
        diff_cell = ws.cell(row=row_idx, column=10)
        diff_cell.value         = f"=I{row_idx}-H{row_idx}"
        diff_cell.number_format = QTY_FMT
        diff_cell.alignment     = Alignment(horizontal="center", vertical="center")

    # ── Formatação condicional na coluna Diferença ─────────────────
    last_row = len(rows) + 1
    if last_row > 1:
        ws.conditional_formatting.add(
            f"J2:J{last_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFCDD2")),
        )
        ws.conditional_formatting.add(
            f"J2:J{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C8E6C9")),
        )

    # ── Aba de legenda ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Legenda")
    legenda = [
        ["Campo",           "Descrição"],
        ["Código",          "SKU / código do produto no Bling"],
        ["Preço Custo",     "Preço de custo registrado no Bling (sem markup)"],
        ["Preço Venda",     "Preço de venda atual no Bling"],
        ["Estoque Bling",   "Saldo físico total no Bling (pode estar desatualizado)"],
        ["Contagem Física", "Preencha aqui a quantidade física contada na loja (fundo amarelo)"],
        ["Diferença",       "=Contagem Física − Estoque Bling  (verde=sobra, vermelho=falta)"],
        [],
        ["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Total de itens", len(rows)],
    ]
    for r, row_data in enumerate(legenda, start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = BOLD_FONT
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 60

    wb.save(output_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta estoque Bling para Excel")
    parser.add_argument("--token", metavar="ACCESS_TOKEN",
                        help="Access token Bling (use quando o refresh_token expirou)")
    args = parser.parse_args()

    print("=" * 60)
    print("  Exportar Estoque Bling → Excel — Papelaria Bibelô")
    print("=" * 60)

    # 1. Carrega credenciais
    if not ENV_FILE.exists():
        sys.exit(f"❌  .env não encontrado em {ENV_FILE}")
    env = load_env(ENV_FILE)

    # 2. Obtém token (via argumento ou banco de dados)
    if args.token:
        print("\n[1/5] Usando token passado via --token")
        print("[2/5] Pulando verificação de banco (modo --token)")
        token = args.token
        conn  = None
    else:
        print("\n[1/5] Conectando ao banco de dados...")
        try:
            conn = get_db_conn(env)
        except Exception as e:
            print(f"\n❌  Falha ao conectar no banco: {e}")
            print("\nO container bibelo_postgres precisa estar rodando.")
            print("Execute: cd /opt/pessoal/bibelocrm && docker compose up -d postgres")
            print("\nOu passe um access_token diretamente:")
            print("  python3 scripts/exportar_estoque_excel.py --token SEU_TOKEN")
            sys.exit(1)

        print("[2/5] Verificando token Bling...")
        try:
            token = get_valid_token(env, conn)
        except Exception as e:
            print(f"\n❌  {e}")
            print("\nO refresh_token provavelmente expirou.")
            print("Para resolver, faça o re-login no BibelôCRM:")
            print("  cd /opt/pessoal/bibelocrm && docker compose up -d api redis")
            print("  Acesse: https://crm.papelariabibelo.com.br → Configurações → Bling → Conectar")
            print("\nOu passe um access_token diretamente:")
            print("  python3 scripts/exportar_estoque_excel.py --token SEU_TOKEN")
            sys.exit(1)

    # 4. Busca categorias (para resolver nomes)
    print("[3/5] Buscando categorias...")
    categories = fetch_categories(token)
    print(f"  {len(categories)} categorias carregadas")

    # 5. Busca produtos
    print("[4/5] Buscando produtos...")
    raw_products = fetch_all_products(token)

    if not raw_products:
        conn.close()
        sys.exit("⚠️  Nenhum produto retornado pela API do Bling.")

    # 6. Busca estoque
    print("[5/5] Buscando saldos de estoque...")
    product_ids = [str(p["id"]) for p in raw_products]
    saldos = fetch_stock_for_ids(product_ids, token)

    # 7. Monta lista consolidada (somente ativos com estoque > 0)
    rows = []
    sem_estoque = 0
    for prod in raw_products:
        ativo = prod.get("situacao") in ("A", "Ativo", True)
        if not ativo:
            continue
        pid   = str(prod.get("id", ""))
        saldo = saldos.get(pid, 0)
        if saldo <= 0:
            sem_estoque += 1
            continue

        cat_id  = (prod.get("categoria") or {}).get("id")
        cat_nome = categories.get(cat_id, "—") if cat_id else "—"

        rows.append({
            "codigo":      prod.get("codigo") or "",
            "nome":        prod.get("nome") or "Sem nome",
            "categoria":   cat_nome,
            "tipo":        prod.get("tipo") or "P",
            "unidade":     prod.get("unidade") or "UN",
            "preco_custo": float(prod.get("precoCusto") or 0),
            "preco_venda": float(prod.get("preco") or 0),
            "saldo_fisico": saldo,
        })

    rows.sort(key=lambda r: (r["categoria"], r["nome"]))

    print(f"\n  ✓  {len(rows)} produtos com estoque")
    print(f"  ·  {sem_estoque} produtos sem estoque (ignorados)")

    # 8. Gera Excel
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = SCRIPT_DIR.parent / f"estoque_bibelo_{timestamp}.xlsx"
    print(f"\n  Gerando Excel em {output_path.name}...")
    build_excel(rows, output_path)

    if conn:
        conn.close()

    print(f"\n✅  Arquivo gerado: {output_path}")
    print(f"   {len(rows)} produtos | Preencha a coluna 'Contagem Física' (amarela)")
    print("=" * 60)


if __name__ == "__main__":
    main()
